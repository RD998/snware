import openpyxl as xl # type: ignore
def search(xlsx, sheetnm, ir, tim, cont):
    og = xl.load_workbook(filename=xlsx)
    
    wrt = og[f'{sheetnm}']
    
    rng = []
    rf = []
    cf = []



    for dol in wrt.iter_cols(max_col= 24, min_col= 2, max_row= 2, min_row=2 ):
        for cil in dol:
            if cil.value == ir :
                a = int(cil.column) + 4
                rng = [int(cil.column), a]

    for col in wrt.iter_cols(max_col= int(rng[1]), min_col= int(rng[0]), max_row= 3, min_row=3 ):
        for cell in col:
            if str(cell.value) == tim :
                cf = cell.column_letter

    for ro in wrt.iter_rows(max_col= 1, min_col= 1, max_row= 24, min_row=5 ):
        for op in ro:
            if str(op.value) == cont :
                rf = op.row
                
    f = wrt[f'{str(cf) + str(rf)}']

    return (f'${f.value}')
            

def quota(file, cont:int, ban:int, weight:int, lvl, loi ):
    wrk = xl.load_workbook(filename= file)
    wrks = wrk['SNWare']
    lookupVAL = lvl + loi
    for element in wrks.iter_rows(max_col=9, min_col=9, min_row=2, max_row=49):
        for cell in element:
            if str(cell.value) == lookupVAL:
                TATtime = wrks[f'J{cell.row}'].value
                TATdays = wrks[f'K{cell.row}'].value
                baseCHRG = wrks[f'M{cell.row}'].value
                chngHrs = wrks[f'L{cell.row}'].value
        
        for nvr in wrks.iter_rows(max_col=18, min_col=18, max_row=12, min_row=3):
            for never in nvr:
                if str(never.value) == loi:
                    AddCont = wrks[f'X{never.row}'].value
                    AddBan = wrks[f'Y{never.row}'].value
                    AddWeight = wrks[f'Z{never.row}'].value
    
    final_cost = baseCHRG + ((cont - 1)*AddCont) + ((ban -1)*AddBan) + ((weight)*(AddWeight))

    return final_cost



