import openpyxl as xl

def quota(cont:int, ban:int, weight:int, lvl, loi )
    wrk = xl.load_workbook(filename= 'Snware_DP_Quote.xlsx')
    wrks = wrk['SNWare']
    lookupVAL = lvl + loi
    for element in wrks.iter_rows(max_col=9, min_col=9, min_row=2, max_row=49):
        for cell in element:
            if str(cell.value) == lookupVAL:
                TATtime = wrks[f'J{cell.row}'].value
                TATdays = wrks[f'K{cell.row}'].value
                baseCHRG = wrks[f'M{cell.row}'].value
                chngHrs = wrks[f'L{cell.row}'].value
        
        for nvr in wrks.iter_rows(max_col=18, min_col=18, max_row=12, min_row=3):
            for never in nvr:
                if str(never.value) == loi:
                    AddCont = wrks[f'X{never.row}'].value
                    AddBan = wrks[f'Y{never.row}'].value
                    AddWeight = wrks[f'Z{never.row}'].value
    
    final_cost = baseCHRG + ((cont - 1)*AddCont) + ((ban -1)*AddBan) + ((weight)*(AddWeight))

    return final_cost
                
    